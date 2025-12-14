from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Sequence

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook


# -----------------------------
# Visual palette (dark-safe, high contrast)
# -----------------------------
BLOCK_COLORS = [
    "#00B5F7",  # electric blue
    "#00E676",  # neon green
    "#FF6D00",  # vivid orange
    "#E040FB",  # neon purple
    "#FF1744",  # vivid red
]
PORTFOLIO_COLOR = "#FFD700"  # gold


# -----------------------------
# Primitives
# -----------------------------
@dataclass(frozen=True)
class CellRef:
    value: str


@dataclass(frozen=True)
class LegId:
    value: str


@dataclass(frozen=True)
class BlockId:
    value: str


@dataclass(frozen=True)
class Strike:
    value: float


@dataclass(frozen=True)
class PremiumPerArroba:
    value: float

    def abs(self) -> "PremiumPerArroba":
        return PremiumPerArroba(abs(self.value))


@dataclass(frozen=True)
class Contracts:
    value: float

    def abs(self) -> float:
        return float(abs(self.value))


@dataclass(frozen=True)
class PriceGrid:
    values: np.ndarray

    @staticmethod
    def from_range(min_price: float, max_price: float, points: int) -> "PriceGrid":
        return PriceGrid(np.linspace(min_price, max_price, points))


@dataclass(frozen=True)
class PayoffCurve:
    values: np.ndarray

    def plus(self, other: "PayoffCurve") -> "PayoffCurve":
        return PayoffCurve(self.values + other.values)

    def scaled(self, factor: float) -> "PayoffCurve":
        return PayoffCurve(self.values * float(factor))


# -----------------------------
# Instruments
# -----------------------------
@dataclass(frozen=True)
class InstrumentKind:
    value: str  # "call" | "put" | "future"

    def is_put(self) -> bool:
        return self.value == "put"

    def is_future(self) -> bool:
        return self.value == "future"


@dataclass(frozen=True)
class PositionKind:
    value: str  # "short" | "long"

    @staticmethod
    def from_contracts(contracts: Contracts) -> "PositionKind":
        if contracts.value < 0.0:
            return PositionKind("short")
        return PositionKind("long")


@dataclass(frozen=True)
class InstrumentLeg:
    leg_id: LegId
    label: str
    kind: InstrumentKind
    strike: Strike  # for future: entry price
    premium: PremiumPerArroba  # for future: 0
    contracts: Contracts


@dataclass(frozen=True)
class InstrumentBlock:
    block_id: BlockId
    label: str
    legs: tuple[InstrumentLeg, ...]

    def strikes(self) -> list[float]:
        return [leg.strike.value for leg in self.legs]


# -----------------------------
# Excel access
# -----------------------------
class Worksheet:
    def __init__(self, sheet) -> None:
        self._sheet = sheet

    def float_at(self, ref: CellRef) -> float:
        value = self._sheet[ref.value].value
        if value is None:
            raise ValueError(f"Célula vazia: {ref.value}")
        return float(value)

    def optional_float_at(self, ref: CellRef) -> float | None:
        value = self._sheet[ref.value].value
        if value is None:
            return None
        return float(value)

    def has_number_at(self, ref: CellRef) -> bool:
        return self._sheet[ref.value].value is not None

    def value_at(self, addr: str):
        return self._sheet[addr].value

    def iter_cells(self, row_min: int, row_max: int, col_min: int, col_max: int):
        for row in self._sheet.iter_rows(
            min_row=row_min, max_row=row_max, min_col=col_min, max_col=col_max
        ):
            for cell in row:
                yield cell


class WorkbookReader:
    def __init__(self, path: Path, sheet_name: str) -> None:
        self._path = path
        self._sheet_name = sheet_name

    def load_sheet(self) -> Worksheet:
        workbook = load_workbook(self._path, data_only=True)
        sheet = workbook[self._sheet_name]
        return Worksheet(sheet)


# -----------------------------
# Discovery (options/futures)
# -----------------------------
@dataclass(frozen=True)
class LegCandidate:
    leg_id: LegId
    label: str
    kind: InstrumentKind
    value_col: str
    strike_cell: CellRef
    premium_cell: CellRef
    contracts_cell: CellRef

    def is_selectable(self, ws: Worksheet) -> bool:
        if self.kind.is_future():
            return ws.has_number_at(self.strike_cell) and ws.has_number_at(self.contracts_cell)
        return (
            ws.has_number_at(self.strike_cell)
            and ws.has_number_at(self.premium_cell)
            and ws.has_number_at(self.contracts_cell)
        )

    def to_leg(self, ws: Worksheet) -> InstrumentLeg:
        strike = Strike(ws.float_at(self.strike_cell))
        contracts = Contracts(ws.float_at(self.contracts_cell))

        if self.kind.is_future():
            return InstrumentLeg(
                leg_id=self.leg_id,
                label=self.label,
                kind=self.kind,
                strike=strike,
                premium=PremiumPerArroba(0.0),
                contracts=contracts,
            )

        premium = PremiumPerArroba(ws.float_at(self.premium_cell))
        return InstrumentLeg(
            leg_id=self.leg_id,
            label=self.label,
            kind=self.kind,
            strike=strike,
            premium=premium,
            contracts=contracts,
        )


class BaseLegFinder:
    _pattern = re.compile(r"^(Call|Put)[_\s\-].+", re.IGNORECASE)

    def __init__(self, ws: Worksheet) -> None:
        self._ws = ws

    def find(self) -> list[LegCandidate]:
        candidates: list[LegCandidate] = []
        for cell in self._ws.iter_cells(row_min=1, row_max=1, col_min=1, col_max=500):
            if not isinstance(cell.value, str):
                continue
            label = cell.value.strip()
            if not self._pattern.match(label):
                continue
            kind = self._infer_kind(label)
            candidates.append(self._candidate_from_label_cell(cell.coordinate, label, kind))
        return candidates

    def _infer_kind(self, label: str) -> InstrumentKind:
        if label.strip().upper().startswith("CALL"):
            return InstrumentKind("call")
        return InstrumentKind("put")

    def _candidate_from_label_cell(self, coord: str, label: str, kind: InstrumentKind) -> LegCandidate:
        col_letters = re.findall(r"[A-Z]+", coord)[0]
        value_col_index = self._col_to_int(col_letters) + 1
        value_col_letters = self._int_to_col(value_col_index)

        leg_id = LegId(f"{label}::{value_col_letters}")
        return LegCandidate(
            leg_id=leg_id,
            label=label,
            kind=kind,
            value_col=value_col_letters,
            strike_cell=CellRef(f"{value_col_letters}1"),
            premium_cell=CellRef(f"{value_col_letters}14"),
            contracts_cell=CellRef(f"{value_col_letters}7"),
        )

    def _col_to_int(self, col: str) -> int:
        result = 0
        for ch in col:
            result = result * 26 + (ord(ch) - ord("A") + 1)
        return result

    def _int_to_col(self, n: int) -> str:
        letters: list[str] = []
        while n > 0:
            n, rem = divmod(n - 1, 26)
            letters.append(chr(rem + ord("A")))
        return "".join(reversed(letters))


class FutureBlotterFinder:
    """
    Reads futures from Base blotter, cols V:AF.

    Mapping:
      Dez_25: qty V, price W
      Jan_26: qty Y, price Z
      Fev_26: qty AB, price AC
      Mar_26: qty AE, price AF

    Rows:
      month label: row 1
      avg price: row 27
      net futures qty: row 37
    """

    def __init__(self, ws: Worksheet) -> None:
        self._ws = ws

    def find(self) -> list[LegCandidate]:
        return [self._candidate_for_month(spec) for spec in self._month_specs()]

    def _month_specs(self) -> list[dict[str, str]]:
        return [
            {"month_cell": "V1", "qty_col": "V", "px_col": "W", "px_row": "27", "qty_row": "37"},
            {"month_cell": "Y1", "qty_col": "Y", "px_col": "Z", "px_row": "27", "qty_row": "37"},
            {"month_cell": "AB1", "qty_col": "AB", "px_col": "AC", "px_row": "27", "qty_row": "37"},
            {"month_cell": "AE1", "qty_col": "AE", "px_col": "AF", "px_row": "27", "qty_row": "37"},
        ]

    def _candidate_for_month(self, spec: dict[str, str]) -> LegCandidate:
        month = self._read_month_label(spec["month_cell"])
        label = f"Futuro_{month}"
        leg_id = LegId(f"{label}::{spec['qty_col']}")

        return LegCandidate(
            leg_id=leg_id,
            label=label,
            kind=InstrumentKind("future"),
            value_col=spec["qty_col"],
            strike_cell=CellRef(f"{spec['px_col']}{spec['px_row']}"),
            premium_cell=CellRef(f"{spec['px_col']}{spec['px_row']}"),
            contracts_cell=CellRef(f"{spec['qty_col']}{spec['qty_row']}"),
        )

    def _read_month_label(self, addr: str) -> str:
        value = self._ws.value_at(addr)
        if isinstance(value, str) and value.strip():
            return value.strip()
        return addr


# -----------------------------
# Catalog (UI-friendly)
# -----------------------------
class LegCatalog:
    def __init__(self, ws: Worksheet, candidates: Sequence[LegCandidate]) -> None:
        self._ws = ws
        self._candidates = list(candidates)
        self._by_id = {c.leg_id.value: c for c in candidates}
        self._display_cache: dict[str, str] = {}

    def filtered_ids(self, text: str, kind: str) -> list[str]:
        needle = text.strip().lower()
        ids: list[str] = []
        for c in self._candidates:
            if kind != "Todos" and c.kind.value != kind:
                continue
            if needle and needle not in c.label.lower() and needle not in c.value_col.lower():
                continue
            ids.append(c.leg_id.value)
        return ids

    def display_name(self, leg_id: str) -> str:
        cached = self._display_cache.get(leg_id)
        if cached is not None:
            return cached

        c = self._by_id[leg_id]
        name = self._display_future(c) if c.kind.is_future() else self._display_option(c)
        self._display_cache[leg_id] = name
        return name

    def to_leg(self, leg_id: str) -> InstrumentLeg:
        return self._by_id[leg_id].to_leg(self._ws)

    def _display_future(self, c: LegCandidate) -> str:
        entry = self._ws.optional_float_at(c.strike_cell)
        cts = self._ws.optional_float_at(c.contracts_cell)
        e_txt = "NA" if entry is None else f"{entry:.2f}"
        c_txt = "NA" if cts is None else f"{cts:.2f}"
        return f"{c.label} | entry={e_txt} | cts={c_txt}"

    def _display_option(self, c: LegCandidate) -> str:
        strike = self._ws.optional_float_at(c.strike_cell)
        prem = self._ws.optional_float_at(c.premium_cell)
        cts = self._ws.optional_float_at(c.contracts_cell)

        k_txt = "NA" if strike is None else f"{strike:.2f}"
        p_txt = "NA" if prem is None else f"{prem:.4f}"
        c_txt = "NA" if cts is None else f"{cts:.2f}"
        return f"{c.label} | col {c.value_col} | K={k_txt} | prem={p_txt} | cts={c_txt}"


# -----------------------------
# Payoff engine
# -----------------------------
class PayoffModel:
    def __init__(self, grid: PriceGrid) -> None:
        self._s = grid.values

    def curve_for_leg(self, leg: InstrumentLeg) -> PayoffCurve:
        if leg.kind.is_future():
            return PayoffCurve(self._curve_future(leg.strike.value, leg.contracts))

        position = PositionKind.from_contracts(leg.contracts)
        premium = leg.premium.abs().value

        curve = self._curve_call(leg.strike.value, premium, position)
        if leg.kind.is_put():
            curve = self._curve_put(leg.strike.value, premium, position)
        return PayoffCurve(curve)

    def _curve_call(self, k: float, p: float, pos: PositionKind) -> np.ndarray:
        intrinsic = np.maximum(self._s - k, 0.0)
        if pos.value == "short":
            return p - intrinsic
        return intrinsic - p

    def _curve_put(self, k: float, p: float, pos: PositionKind) -> np.ndarray:
        intrinsic = np.maximum(k - self._s, 0.0)
        if pos.value == "short":
            return p - intrinsic
        return intrinsic - p

    def _curve_future(self, entry: float, contracts: Contracts) -> np.ndarray:
        if contracts.value < 0.0:
            return entry - self._s
        return self._s - entry


class CurveAggregator:
    def weighted_average(self, curves: Sequence[PayoffCurve], weights: Sequence[float]) -> PayoffCurve:
        matrix = np.vstack([c.values for c in curves])
        w = np.asarray(weights, dtype=float).reshape(-1, 1)
        total = float(w.sum())
        if total == 0.0:
            return PayoffCurve(np.zeros(matrix.shape[1]))
        return PayoffCurve((matrix * w).sum(axis=0) / total)


# -----------------------------
# Results for analytics
# -----------------------------
@dataclass(frozen=True)
class LegResult:
    label: str
    kind: str
    strike_or_entry: float
    premium: float
    contracts: float
    curve: PayoffCurve


@dataclass(frozen=True)
class BlockResult:
    block: InstrumentBlock
    legs: tuple[LegResult, ...]
    per_arroba: PayoffCurve
    total_reais: PayoffCurve
    abs_contracts_sum: float


# -----------------------------
# Plot configuration
# -----------------------------
@dataclass(frozen=True)
class PlotConfig:
    mode: str  # "Apresentação" | "Executivo" | "Analítico"
    show_total: bool
    show_legs: bool
    show_strikes: bool
    blocks_with_legs: tuple[str, ...]
    hide_aggregates_when_legs: bool


class PlotBuilder:
    def __init__(self, grid: PriceGrid) -> None:
        self._x = grid.values

    def payoff_per_arroba(
        self,
        results: Sequence[BlockResult],
        portfolio: PayoffCurve,
        strikes: Sequence[float],
        config: PlotConfig,
    ) -> go.Figure:
        fig = go.Figure()

        # 1) legs first (only in analytic)
        if config.show_legs:
            for idx, r in enumerate(results):
                base_color = BLOCK_COLORS[idx % len(BLOCK_COLORS)]
                if r.block.label not in config.blocks_with_legs:
                    continue
                for leg in r.legs:
                    fig.add_trace(
                        self._line(
                            curve=leg.curve,
                            name=f"{r.block.label} :: {leg.label}",
                            width=1,
                            opacity=0.25,
                            color=base_color,
                            dash="dot",
                        )
                    )

        # 2) aggregates next (presentation/executive/analytic)
        for idx, r in enumerate(results):
            base_color = BLOCK_COLORS[idx % len(BLOCK_COLORS)]
            fig.add_trace(
                self._line(
                    curve=r.per_arroba,
                    name=f"{r.block.label} (agregado)",
                    width=4,
                    opacity=1.0,
                    color=base_color,
                    dash="solid",
                )
            )

        # 3) portfolio last (always on top)
        fig.add_trace(
            self._line(
                curve=portfolio,
                name="PORTFÓLIO (R$/@)",
                width=6,
                opacity=1.0,
                color=PORTFOLIO_COLOR,
                dash="solid",
            )
        )

        if config.show_strikes:
            for k in sorted(set(strikes)):
                fig.add_vline(x=float(k), line_width=1, line_dash="dash", opacity=0.35)

        fig.add_hline(y=0, line_width=1, line_dash="dash", opacity=0.6)
        fig.update_layout(
            title="Payoff por @ (R$/@) – blocos selecionados",
            xaxis_title="Preço (R$/@)",
            yaxis_title="Payoff (R$/@)",
            hovermode="x unified",
            legend_title="Curvas",
            legend=dict(bgcolor="rgba(0,0,0,0)", borderwidth=0, font=dict(size=12)),
        )
        return fig

    def payoff_total_reais(
        self,
        results: Sequence[BlockResult],
        portfolio_total: PayoffCurve,
        strikes: Sequence[float],
        config: PlotConfig,
    ) -> go.Figure:
        fig = go.Figure()

        for idx, r in enumerate(results):
            base_color = BLOCK_COLORS[idx % len(BLOCK_COLORS)]
            fig.add_trace(
                self._line(
                    curve=r.total_reais,
                    name=f"{r.block.label} (R$)",
                    width=4,
                    opacity=1.0,
                    color=base_color,
                    dash="solid",
                )
            )

        fig.add_trace(
            self._line(
                curve=portfolio_total,
                name="PORTFÓLIO (R$)",
                width=6,
                opacity=1.0,
                color=PORTFOLIO_COLOR,
                dash="solid",
            )
        )

        if config.show_strikes:
            for k in sorted(set(strikes)):
                fig.add_vline(x=float(k), line_width=1, line_dash="dash", opacity=0.35)

        fig.add_hline(y=0, line_width=1, line_dash="dash", opacity=0.6)
        fig.update_layout(
            title="Payoff total aproximado (R$) – blocos selecionados",
            xaxis_title="Preço (R$/@)",
            yaxis_title="Payoff (R$)",
            hovermode="x unified",
            legend_title="Curvas",
            legend=dict(bgcolor="rgba(0,0,0,0)", borderwidth=0, font=dict(size=12)),
        )
        return fig

    def _line(
        self,
        curve: PayoffCurve,
        name: str,
        width: int,
        opacity: float,
        color: str,
        dash: str = "solid",
    ) -> go.Scatter:
        return go.Scatter(
            x=self._x,
            y=curve.values,
            mode="lines",
            name=name,
            line={"width": width, "color": color, "dash": dash},
            opacity=float(opacity),
        )


# -----------------------------
# Helpers
# -----------------------------
def save_upload(uploaded) -> Path:
    target = Path.cwd() / "uploaded_planilha.xlsx"
    target.write_bytes(uploaded.getbuffer())
    return target


def build_blocks(selection: Mapping[str, list[str]], catalog: LegCatalog) -> list[InstrumentBlock]:
    blocks: list[InstrumentBlock] = []
    for label, leg_ids in selection.items():
        legs = [catalog.to_leg(leg_id) for leg_id in leg_ids]
        blocks.append(InstrumentBlock(BlockId(label), label, tuple(legs)))
    return blocks


def short_label(text: str, max_len: int = 32) -> str:
    t = text.strip()
    if len(t) <= max_len:
        return t
    return f"{t[: max_len - 3]}..."


def plot_config(mode: str, show_total: bool, block_labels: Sequence[str]) -> PlotConfig:
    if mode == "Apresentação":
        return PlotConfig(
            mode=mode,
            show_total=show_total,
            show_legs=False,
            show_strikes=False,
            blocks_with_legs=tuple(),
            hide_aggregates_when_legs=True,
        )

    if mode == "Executivo":
        return PlotConfig(
            mode=mode,
            show_total=show_total,
            show_legs=False,
            show_strikes=False,
            blocks_with_legs=tuple(),
            hide_aggregates_when_legs=True,
        )

    default = list(block_labels[:1]) if block_labels else []
    blocks_with_legs = st.sidebar.multiselect(
        "Blocos para detalhar (mostrar legs)",
        options=list(block_labels),
        default=default,
    )
    return PlotConfig(
        mode=mode,
        show_total=show_total,
        show_legs=True,
        show_strikes=True,
        blocks_with_legs=tuple(blocks_with_legs),
        hide_aggregates_when_legs=False,
    )


def analytic_summary_table(
    results: Sequence[BlockResult],
    x_grid: np.ndarray,
    s_min: float,
    s_max: float,
    arrobas_per_contract: float,
) -> pd.DataFrame:
    s_mid = (s_min + s_max) / 2.0
    rows: list[dict[str, float | str]] = []

    for r in results:
        contracts_sum = float(sum(l.contracts for l in r.legs))
        abs_contracts_sum = float(sum(abs(l.contracts) for l in r.legs))

        rows.append(
            {
                "Nível": "BLOCO",
                "Bloco": r.block.label,
                "Leg": "",
                "Tipo": "",
                "Strike/Entry": "",
                "Prêmio (R$/@)": "",
                "Contratos": contracts_sum,
                "|Contratos|": abs_contracts_sum,
                "Payoff (R$/@) @Smin": float(np.interp(s_min, x_grid, r.per_arroba.values)),
                "Payoff (R$/@) @Smid": float(np.interp(s_mid, x_grid, r.per_arroba.values)),
                "Payoff (R$/@) @Smax": float(np.interp(s_max, x_grid, r.per_arroba.values)),
                "Payoff (R$) @Smin": float(np.interp(s_min, x_grid, r.total_reais.values)),
                "Payoff (R$) @Smid": float(np.interp(s_mid, x_grid, r.total_reais.values)),
                "Payoff (R$) @Smax": float(np.interp(s_max, x_grid, r.total_reais.values)),
            }
        )

        for leg in r.legs:
            abs_cts = abs(float(leg.contracts))
            total_curve = leg.curve.scaled(abs_cts * float(arrobas_per_contract))

            rows.append(
                {
                    "Nível": "LEG",
                    "Bloco": r.block.label,
                    "Leg": leg.label,
                    "Tipo": leg.kind,
                    "Strike/Entry": float(leg.strike_or_entry),
                    "Prêmio (R$/@)": float(leg.premium),
                    "Contratos": float(leg.contracts),
                    "|Contratos|": float(abs_cts),
                    "Payoff (R$/@) @Smin": float(np.interp(s_min, x_grid, leg.curve.values)),
                    "Payoff (R$/@) @Smid": float(np.interp(s_mid, x_grid, leg.curve.values)),
                    "Payoff (R$/@) @Smax": float(np.interp(s_max, x_grid, leg.curve.values)),
                    "Payoff (R$) @Smin": float(np.interp(s_min, x_grid, total_curve.values)),
                    "Payoff (R$) @Smid": float(np.interp(s_mid, x_grid, total_curve.values)),
                    "Payoff (R$) @Smax": float(np.interp(s_max, x_grid, total_curve.values)),
                }
            )

    df = pd.DataFrame(rows)
    cols_first = ["Nível", "Bloco", "Leg", "Tipo", "Strike/Entry", "Prêmio (R$/@)", "Contratos", "|Contratos|"]
    cols_rest = [c for c in df.columns if c not in cols_first]
    return df[cols_first + cols_rest]


# -----------------------------
# App
# -----------------------------
def main() -> None:
    st.set_page_config(layout="wide")
    st.title("Payoff – Blocos (opções + futuros do blotter)")

    uploaded = st.file_uploader("Carregue o arquivo Excel (.xlsx)", type=["xlsx"])
    if uploaded is None:
        st.stop()

    excel_path = save_upload(uploaded)

    sheet_name = st.sidebar.text_input("Aba (sheet) para ler", value="Base")
    ws = WorkbookReader(excel_path, sheet_name).load_sheet()

    option_candidates = BaseLegFinder(ws).find()
    option_candidates = [c for c in option_candidates if c.is_selectable(ws)]

    include_futures = st.sidebar.checkbox("Incluir futuros do blotter", value=True)
    future_candidates: list[LegCandidate] = []
    if include_futures:
        future_candidates = FutureBlotterFinder(ws).find()
        future_candidates = [c for c in future_candidates if c.is_selectable(ws)]

    candidates = option_candidates + future_candidates
    if len(candidates) == 0:
        st.error("Não encontrei instrumentos válidos na aba Base.")
        st.stop()

    catalog = LegCatalog(ws, candidates)

    st.sidebar.subheader("Seleção de blocos")

    if "blocks" not in st.session_state:
        st.session_state["blocks"] = {"Bloco 1": []}

    if st.sidebar.button("Adicionar bloco"):
        st.session_state["blocks"][f"Bloco {len(st.session_state['blocks']) + 1}"] = []

    block_names = list(st.session_state["blocks"].keys())
    active_block = st.sidebar.selectbox("Bloco para editar", options=block_names)

    filter_text = st.sidebar.text_input("Filtrar legs (texto)", value="")
    filter_kind = st.sidebar.selectbox("Tipo", options=["Todos", "call", "put", "future"], index=0)

    available_ids = catalog.filtered_ids(filter_text, filter_kind)
    current = st.session_state["blocks"][active_block]

    chosen = st.sidebar.multiselect(
        "Selecione as pernas deste bloco",
        options=available_ids,
        default=[x for x in current if x in available_ids] or current,
        format_func=catalog.display_name,
    )
    st.session_state["blocks"][active_block] = chosen

    rename_from = st.sidebar.selectbox("Renomear bloco", options=block_names)
    rename_to = st.sidebar.text_input("Novo nome", value=rename_from)
    if st.sidebar.button("Aplicar rename"):
        name = rename_to.strip()
        if name and name != rename_from:
            st.session_state["blocks"][name] = st.session_state["blocks"].pop(rename_from)

    selection = {k: v for k, v in st.session_state["blocks"].items() if len(v) > 0}
    if len(selection) == 0:
        st.warning("Selecione ao menos 1 perna em algum bloco.")
        st.stop()

    blocks = build_blocks(selection, catalog)
    block_labels = [b.label for b in blocks]

    st.sidebar.subheader("Visual")
    mode = st.sidebar.radio("Modo", options=["Apresentação", "Executivo", "Analítico"], index=0)
    show_total = st.sidebar.checkbox("Mostrar payoff total aproximado (R$)", value=False)
    cfg = plot_config(mode, show_total, block_labels)

    strikes_all = [s for b in blocks for s in b.strikes()]
    k_min = float(min(strikes_all))
    k_max = float(max(strikes_all))

    col1, col2, col3 = st.columns(3)
    s_min = col1.slider(
        "S min (R$/@)",
        min_value=0.0,
        max_value=k_max + 300.0,
        value=max(0.0, k_min - 80.0),
        step=1.0,
    )
    s_max = col2.slider(
        "S max (R$/@)",
        min_value=0.0,
        max_value=k_max + 300.0,
        value=k_max + 80.0,
        step=1.0,
    )
    points = col3.slider("Pontos", min_value=101, max_value=2001, value=401, step=50)

    if s_max < s_min:
        s_min, s_max = s_max, s_min

    col4, col5, col6 = st.columns(3)
    premium_mult = col4.slider("Multiplicador de prêmio (opções)", min_value=0.0, max_value=2.0, value=1.0, step=0.01)
    contracts_mult = col5.slider("Multiplicador de contratos (todos)", min_value=0.0, max_value=2.0, value=1.0, step=0.01)
    arrobas_per_contract = col6.number_input("@ por contrato", min_value=1.0, max_value=2000.0, value=330.0, step=1.0)

    grid = PriceGrid.from_range(float(s_min), float(s_max), int(points))
    model = PayoffModel(grid)
    aggregator = CurveAggregator()

    results: list[BlockResult] = []
    portfolio = PayoffCurve(np.zeros_like(grid.values))

    for block in blocks:
        curves: list[PayoffCurve] = []
        weights: list[float] = []
        abs_contracts_sum = 0.0
        leg_results: list[LegResult] = []

        for leg in block.legs:
            contracts = Contracts(leg.contracts.value * float(contracts_mult))

            premium = leg.premium
            if not leg.kind.is_future():
                premium = PremiumPerArroba(leg.premium.abs().value * float(premium_mult))

            adjusted = InstrumentLeg(leg.leg_id, leg.label, leg.kind, leg.strike, premium, contracts)
            curve = model.curve_for_leg(adjusted)

            leg_results.append(
                LegResult(
                    label=short_label(adjusted.label),
                    kind=adjusted.kind.value,
                    strike_or_entry=float(adjusted.strike.value),
                    premium=float(adjusted.premium.value),
                    contracts=float(adjusted.contracts.value),
                    curve=curve,
                )
            )

            curves.append(curve)
            w = adjusted.contracts.abs()
            weights.append(w)
            abs_contracts_sum += w

        per_arroba = aggregator.weighted_average(curves, weights)
        portfolio = portfolio.plus(per_arroba)

        total_reais = per_arroba.scaled(abs_contracts_sum * float(arrobas_per_contract))
        results.append(BlockResult(block, tuple(leg_results), per_arroba, total_reais, abs_contracts_sum))

    plotter = PlotBuilder(grid)

    st.plotly_chart(
        plotter.payoff_per_arroba(results, portfolio, strikes_all, cfg),
        use_container_width=True,
    )

    if cfg.show_total:
        portfolio_total = PayoffCurve(np.zeros_like(grid.values))
        for r in results:
            portfolio_total = portfolio_total.plus(r.total_reais)

        st.plotly_chart(
            plotter.payoff_total_reais(results, portfolio_total, strikes_all, cfg),
            use_container_width=True,
        )

    st.subheader("Resumo dos blocos selecionados")
    st.dataframe(
        analytic_summary_table(results, grid.values, float(s_min), float(s_max), float(arrobas_per_contract)),
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
