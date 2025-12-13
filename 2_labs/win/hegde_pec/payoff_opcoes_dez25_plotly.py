import numpy as np
import pandas as pd
from openpyxl import load_workbook
import plotly.graph_objects as go

# ==============================
# CONFIGURAÇÃO
# ==============================
EXCEL_FILE = "Planilha Hedge.xlsx"   # ajuste o caminho se necessário
SHEET = "Base"

ARROBAS_POR_CONTRATO = 330
N_POINTS = 500

# ==============================
# FUNÇÕES AUXILIARES
# ==============================
def read_float(ws, addr):
    v = ws[addr].value
    if v is None:
        raise ValueError(f"Célula {addr} vazia")
    return float(v)

def payoff_option(S, K, premium, option_type, position):
    if option_type == "call":
        intrinsic = np.maximum(S - K, 0)
    elif option_type == "put":
        intrinsic = np.maximum(K - S, 0)
    else:
        raise ValueError("option_type inválido")

    if position == "short":
        return premium - intrinsic
    elif position == "long":
        return intrinsic - premium
    else:
        raise ValueError("position inválida")

def weighted_curve(curves, weights):
    weights = np.abs(np.array(weights))
    curves = np.array(curves)
    return np.sum(curves * weights[:, None], axis=0) / weights.sum()

# ==============================
# LEITURA DA PLANILHA
# ==============================
wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET]

# PUT Dez/25 – legs EB e EH
put_legs = [
    {
        "name": "Put EB",
        "K": read_float(ws, "EB1"),
        "premium": read_float(ws, "EB14"),
        "contracts": read_float(ws, "EB7"),
    },
    {
        "name": "Put EH",
        "K": read_float(ws, "EH1"),
        "premium": read_float(ws, "EH14"),
        "contracts": read_float(ws, "EH7"),
    },
]

# CALL Dez/25 – legs EE e EK
call_legs = [
    {
        "name": "Call EE",
        "K": read_float(ws, "EE1"),
        "premium": read_float(ws, "EE14"),
        "contracts": read_float(ws, "EE7"),
    },
    {
        "name": "Call EK",
        "K": read_float(ws, "EK1"),
        "premium": read_float(ws, "EK14"),
        "contracts": read_float(ws, "EK7"),
    },
]

spot_ref = float(ws["AV2"].value)
all_strikes = [l["K"] for l in put_legs + call_legs]

# ==============================
# DOMÍNIO DE PREÇO
# ==============================
S_min = max(0, min(all_strikes + [spot_ref]) - 80)
S_max = max(all_strikes + [spot_ref]) + 80
S = np.linspace(S_min, S_max, N_POINTS)

# ==============================
# PAYOFF DAS LEGS
# ==============================
put_curves = []
for leg in put_legs:
    pos = "short" if leg["contracts"] < 0 else "long"
    put_curves.append(
        payoff_option(S, leg["K"], leg["premium"], "put", pos)
    )

call_curves = []
for leg in call_legs:
    pos = "short" if leg["contracts"] < 0 else "long"
    call_curves.append(
        payoff_option(S, leg["K"], leg["premium"], "call", pos)
    )

put_weights = [leg["contracts"] for leg in put_legs]
call_weights = [leg["contracts"] for leg in call_legs]

put_agg = weighted_curve(put_curves, put_weights)
call_agg = weighted_curve(call_curves, call_weights)
net_agg = put_agg + call_agg

# ==============================
# GRÁFICO INTERATIVO (PLOTLY)
# ==============================
fig = go.Figure()

# Legs
for i, leg in enumerate(put_legs):
    fig.add_trace(go.Scatter(
        x=S, y=put_curves[i],
        mode="lines",
        name=leg["name"]
    ))

for i, leg in enumerate(call_legs):
    fig.add_trace(go.Scatter(
        x=S, y=call_curves[i],
        mode="lines",
        name=leg["name"]
    ))

# Agregados
fig.add_trace(go.Scatter(
    x=S, y=put_agg,
    mode="lines",
    name="PUT agregado",
    line=dict(width=4)
))

fig.add_trace(go.Scatter(
    x=S, y=call_agg,
    mode="lines",
    name="CALL agregado",
    line=dict(width=4)
))

fig.add_trace(go.Scatter(
    x=S, y=net_agg,
    mode="lines",
    name="NET (PUT + CALL)",
    line=dict(width=5)
))

# Linhas verticais de strike
for k in all_strikes:
    fig.add_vline(x=k, line_dash="dash", opacity=0.4)

fig.add_hline(y=0, line_dash="dash")

fig.update_layout(
    title="Payoff por @ (R$/@) – Opções Dez/25",
    xaxis_title="Preço do boi (R$/@)",
    yaxis_title="Payoff (R$/@)",
    hovermode="x unified"
)

fig.show()
