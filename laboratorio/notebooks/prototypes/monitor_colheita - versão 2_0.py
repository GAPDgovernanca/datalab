"""
Programa de Análise de Vistorias

Este programa realiza a análise de dados de vistorias, gerando gráficos de barras empilhadas e gráficos de violino com base nos dados fornecidos em um arquivo CSV.
As configurações, como caminho do arquivo, variáveis para os gráficos e diretório de saída, são especificadas em um arquivo `monitor_colheita.yaml`.

Funcionalidades:
- Carregamento de dados de um arquivo CSV.
- Geração de gráficos de barras empilhadas por líder para variáveis especificadas.
- Geração de gráficos de violino para variáveis especificadas, com eixos invertidos e escala de 1 a 5.
- Salvamento dos gráficos gerados em um diretório de saída especificado.
- Salvamento de uma cópia do arquivo de dados original em formato Excel, com formatação condicional aplicada.

Como usar:
1. Prepare um arquivo `monitor_colheita.yaml` com os seguintes campos:
   file_path: 'caminho/para/o/arquivo.csv'
   stacked_chart_variables:
     - 'Variável1'
     - 'Variável2'
     - 'Variável3'
   violin_plot_columns:
     - 'Variável1'
     - 'Variável2'
     - 'Variável3'
   output_folder: 'diretório/de/saída'
   unidade: 'Nome da Unidade'
   periodo: 'Período (formato: DDMM-DDMM)'

2. Execute o programa com o comando:
   python3 monitor_colheita.py --config monitor_colheita.yaml

O programa gerará os gráficos e salvará os resultados no diretório de saída especificado.
"""

import argparse
import logging
import os
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import yaml
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Definição do logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def main():
    parser = argparse.ArgumentParser(description='Survey Response Analysis')
    parser.add_argument('--config', type=str, default='monitor_colheita.yaml', help='Path to the configuration file')
    args = parser.parse_args()
    
    config = load_config(args.config)
    
    data = load_data(config['file_path'])
    data = preprocess_data(data)
    
    # Define o diretório de saída com base na unidade e no período
    output_folder = os.path.join(config['output_folder'], f"{config['unidade']}_{config['periodo']}")
    os.makedirs(output_folder, exist_ok=True)
    
    create_charts(data, output_folder, config['unidade'], config['periodo'], config)
    save_copy_of_source_file_as_excel(data, output_folder, config['unidade'], config['periodo'])
    
    logger.info("Analysis completed successfully")

def load_config(config_path):
    with open(config_path, 'r') as file:
        config = yaml.safe_load(file)
    return config

def load_data(file_path):
    try:
        data = pd.read_csv(file_path)
        logger.info(f"Data loaded successfully from {file_path}")
        return data
    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        raise

def preprocess_data(data):
    # Converter Submit Date (UTC) para horário de Brasília
    data['Submit Date (UTC)'] = pd.to_datetime(data['Submit Date (UTC)']).dt.tz_localize('UTC').dt.tz_convert('America/Sao_Paulo')
    data['Submit Date (UTC)'] = data['Submit Date (UTC)'].dt.tz_localize(None)  # Remove a informação de fuso horário
    data.rename(columns={'Submit Date (UTC)': 'Submit Date (BRT)'}, inplace=True)
    
    # Reposicionar a coluna para ser a primeira
    cols = list(data.columns)
    cols.insert(0, cols.pop(cols.index('Submit Date (BRT)')))
    data = data[cols]
    
    return data

def create_stacked_count_chart(data, output_folder, unidade, periodo, column):
    plt.figure(figsize=(10, 6))
    counts = data.groupby(['Seu nome', column]).size().unstack(fill_value=0)
    ax = counts.plot(kind='bar', stacked=True)
    
    for container in ax.containers:
        labels = [f'{int(v)}' if v > 0 else '' for v in container.datavalues]
        ax.bar_label(container, labels=labels, label_type='center', fontsize=8)
    
    plt.title(f'{column} - vistorias/notas por LÍDER - {periodo}')
    plt.xlabel('Seu nome')
    plt.ylabel('Counts')
    plt.tight_layout()
    
    # Adiciona linhas de grade em cinza claro
    ax.grid(True, which='both', linestyle='--', linewidth=0.5, color='lightgrey')
    ax.set_axisbelow(True)
    
    output_path = os.path.join(output_folder, f'stacked_bar_{column}.png')
    plt.savefig(output_path)
    plt.close()
    logger.info(f"Stacked bar chart for {column} saved at {output_path}")

def filter_numeric_columns(data, numeric_columns):
    return data[numeric_columns]

def create_violin_plot(data, unidade, periodo):
    plt.figure(figsize=(10, 6))
    ax = sns.violinplot(data=data, orient='h')
    ax.set_xticks([1, 2, 3, 4, 5])
    ax.set_xticklabels(['muito ruim', 'ruim', 'médio', 'bom', 'muito bom'])
    
    plt.title(f'{unidade} - NOTAS DE VISTORIA - {periodo}')
    plt.xlabel('Notas')
    plt.ylabel('')
    plt.tight_layout()
    
    # Adiciona linhas de grade em cinza claro
    ax.grid(True, which='both', linestyle='--', linewidth=0.5, color='lightgrey')
    ax.set_axisbelow(True)
    
    return f'violin_plot_{unidade}_{periodo}.png'

def save_plot(filename, output_folder):
    output_path = os.path.join(output_folder, filename)
    plt.savefig(output_path)
    plt.close()
    logger.info(f"Violin plot saved at {output_path}")

def apply_conditional_formatting(ws, columns):
    for col in columns:
        c_range = f"{col}2:{col}{ws.max_row}"
        rule = ColorScaleRule(start_type='num', start_value=1, start_color='FFB3BA',
                              mid_type='num', mid_value=3, mid_color='FFDFBA',
                              end_type='num', end_value=5, end_color='BAE1FF')
        ws.conditional_formatting.add(c_range, rule)
        for row in ws[c_range]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

def save_copy_of_source_file_as_excel(data, output_folder, unidade, periodo):
    columns_to_drop = ['#', 'Response Type', 'Start Date (UTC)', 'Stage Date (UTC)', 'Network ID', 'Tags']
    data = data.drop(columns=columns_to_drop, errors='ignore')
    
    base_name = f'{unidade}_{periodo}'
    new_file_path = os.path.join(output_folder, f'{base_name}.xlsx')
    
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        worksheet.auto_filter.ref = worksheet.dimensions
    
    wb = load_workbook(new_file_path)
    ws = wb.active
    columns = ['PISOTEIO', 'DANOS', 'FOCO']
    column_letters = [get_column_letter(data.columns.get_loc(col) + 1) for col in columns]
    apply_conditional_formatting(ws, column_letters)
    wb.save(new_file_path)
    
    print(f'Cópia do arquivo de origem salva em: {new_file_path}')

def create_charts(data, output_folder, unidade, periodo, config):
    stacked_chart_columns = config['stacked_chart_variables']
    for column in stacked_chart_columns:
        create_stacked_count_chart(data, output_folder, unidade, periodo, column)
    
    numeric_columns = config['violin_plot_columns']
    data_numeric = filter_numeric_columns(data, numeric_columns)
    titulo = create_violin_plot(data_numeric, unidade, periodo)
    save_plot(titulo, output_folder)

if __name__ == "__main__":
    main()