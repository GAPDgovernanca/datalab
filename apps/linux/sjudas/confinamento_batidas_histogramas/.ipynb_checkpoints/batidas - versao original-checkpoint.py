import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator, FuncFormatter
import datetime
from datetime import timedelta
import io
import base64
import pytz
import os
import yaml
import openpyxl
from openpyxl.styles import PatternFill
import textwrap
from typing import List, Dict, Optional

# 1. CONFIGURAÇÃO E INICIALIZAÇÃO


def read_config(config_file="config.yaml"):
    """
    Lê o arquivo de configuração YAML e carrega as configurações para o programa.

    Args:
    config_file (str): Caminho para o arquivo de configuração YAML. Padrão: 'config.yaml'.

    Returns:
    dict: Um dicionário contendo todas as configurações do programa, como colunas do Excel,
    parâmetros de análise, visualização, sliders e exportação.
    """
    # Obter o diretório do script atual
    base_path = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(base_path, config_file)

    try:
        with open(config_path, "r") as file:
            return yaml.safe_load(file)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"O arquivo de configuração '{config_path}' não foi encontrado."
        )
    except yaml.YAMLError as e:
        raise ValueError(f"Erro ao ler o arquivo de configuração: {e}")


config = read_config()

st.set_page_config(
    page_title=config["ui"]["page_title"], layout="wide"
)  # Move para fora do main()

# 2. FUNÇÕES DE PROCESSAMENTO DE DADOS


def preprocess_dataframe(df, config):
    """
    Pré-processa um DataFrame, convertendo colunas para numéricas e verificando colunas obrigatórias.

    Parâmetros:
    df (DataFrame): DataFrame a ser processado.
    config (dict): Configurações com colunas numéricas e obrigatórias.

    Retorna:
    DataFrame: DataFrame processado.

    Exceções:
    ValueError: Se faltar alguma coluna obrigatória.
    """
    numeric_columns = config["numeric_columns"]
    excel_columns = config["excel_columns"]

    required_columns = list(excel_columns.values())
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Colunas ausentes no arquivo: {', '.join(missing_columns)}")

    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def load_and_process_data(uploaded_file):
    """
    Carrega e processa dados de um arquivo Excel.

    Parâmetros:
    uploaded_file: Arquivo Excel (.xlsx) carregado pelo usuário.

    Retorna:
    DataFrame: Dados processados e prontos para análise.
    """
    try:
        analysis_config = config["analysis"]
        excel_columns = config["excel_columns"]
        skip_rows = analysis_config["skip_rows"]
        remove_first_column = analysis_config["remove_first_column"]
        columns_to_remove = analysis_config.get("columns_to_remove", [])

        df = pd.read_excel(uploaded_file, skiprows=skip_rows)

        if remove_first_column:
            df = df.iloc[:, 1:]

        removed_columns = []
        for col in columns_to_remove:
            if col in df.columns:
                df = df.drop(columns=[col])
                removed_columns.append(col)

        df = preprocess_dataframe(df, config)

        return df

    except Exception as e:
        st.error("Erro ao carregar ou processar o arquivo: " f"{str(e)}")
        return None


def calculate_weighted_average_with_weights(df, pesos_relativos, config):
    """
    Calcula a média ponderada para cada batida com base nos pesos relativos dos tipos de alimento,
    usando desvio normalizado e suavização de valores extremos.

    Parâmetros:
    - df (DataFrame): Dados de entrada
    - pesos_relativos (dict): Pesos relativos dos tipos de alimento
    - config (dict): Configurações

    Retorna:
    - DataFrame: Médias ponderadas por batida
    """
    try:
        weighted_config = config["weighted_average"]
        excel_columns = config["excel_columns"]
        previsto_col = weighted_config["previsto_column"]
        diferenca_percentual_col = weighted_config["diferenca_percentual_column"]
        pesos_relativos_col = weighted_config["pesos_relativos_column"]
        cod_batida_col = excel_columns["cod_batida"]

        # Parâmetros de cálculo
        fator_suavizacao = weighted_config["calculo"]["fator_suavizacao"]
        peso_desvio = weighted_config["calculo"]["peso_desvio"]
        min_std = weighted_config["calculo"]["min_std"]

        df_calc = df.copy()

        # Verificações de qualidade
        if df_calc[diferenca_percentual_col].isnull().any():
            st.warning("Existem valores nulos nas diferenças percentuais")
            df_calc = df_calc.dropna(subset=[diferenca_percentual_col])

        if len(df_calc) == 0:
            raise ValueError("Não há dados válidos para calcular a média ponderada")

        # 1. Preparação inicial
        df_calc[diferenca_percentual_col] = df_calc[diferenca_percentual_col].abs()
        df_calc["peso_tipo"] = df_calc[pesos_relativos_col].map(pesos_relativos)

        # 2. Peso pela quantidade
        totals = df_calc.groupby(cod_batida_col)[previsto_col].transform("sum")
        df_calc["peso_quantidade"] = df_calc[previsto_col] / totals

        # 3. Cálculo do desvio normalizado mais suave
        mean_diff = df_calc.groupby(cod_batida_col)[diferenca_percentual_col].transform(
            "mean"
        )
        std_diff = df_calc.groupby(cod_batida_col)[diferenca_percentual_col].transform(
            "std"
        )
        std_diff = std_diff.clip(lower=min_std)  # Evita divisão por zero

        df_calc["desvio_norm"] = (
            (df_calc[diferenca_percentual_col] - mean_diff) / std_diff
        ) * peso_desvio

        # 4. Fator de suavização mais controlado
        df_calc["fator_suavizacao"] = 1 / (
            1 + np.abs(df_calc["desvio_norm"]) / fator_suavizacao
        )

        # 5. Contribuição ponderada final
        df_calc["contribuicao"] = (
            df_calc[diferenca_percentual_col]
            * df_calc["peso_quantidade"]
            * df_calc["peso_tipo"]
            * df_calc["fator_suavizacao"]
        )

        # 6. Agregação final
        result = df_calc.groupby(cod_batida_col).agg(
            peso_total=("peso_quantidade", "sum"), contrib_total=("contribuicao", "sum")
        )

        # 7. Cálculo da média ponderada
        result["MÉDIA PONDERADA (%)"] = (
            result["contrib_total"] / result["peso_total"]
        ).round(2)

        # Validação dos resultados finais
        result["MÉDIA PONDERADA (%)"] = result["MÉDIA PONDERADA (%)"].clip(lower=0)

        if result["MÉDIA PONDERADA (%)"].isnull().any():
            st.warning("Algumas médias não puderam ser calculadas")
            result = result.dropna(subset=["MÉDIA PONDERADA (%)"])

        # 8. Prepara DataFrame final
        return pd.DataFrame(
            {
                cod_batida_col: result.index,
                "MÉDIA PONDERADA (%)": result["MÉDIA PONDERADA (%)"],
            }
        ).reset_index(drop=True)

    except KeyError as e:
        st.error(f"Erro ao acessar coluna: {str(e)}")
        return None
    except Exception as e:
        st.error(f"Erro no cálculo da média ponderada: {str(e)}")
        return None


def find_correct_columns(df, config):
    """
    Encontra os índices corretos das colunas no DataFrame com base no arquivo de configuração.

    Args:
    df (DataFrame): DataFrame contendo os dados carregados.
    config (dict): Dicionário de configuração contendo os nomes das colunas necessárias.

    Returns:
    dict: Um dicionário que mapeia as colunas esperadas para os seus respectivos índices no DataFrame.
    """
    column_indices = {}
    columns = list(df.columns)

    previsto_indices = [
        i for i, col in enumerate(columns) if col == config["excel_columns"]["previsto"]
    ]
    column_indices["previsto"] = previsto_indices[-1]

    column_indices["realizado"] = column_indices["previsto"] + 1
    column_indices["diferenca"] = column_indices["previsto"] + 2

    if (
        columns[column_indices["realizado"]] != config["excel_columns"]["realizado"]
        or columns[column_indices["diferenca"]]
        != config["excel_columns"]["diferenca_percentual"]
    ):
        raise ValueError("A estrutura da planilha não corresponde ao esperado.")

    return column_indices


def remove_outliers_from_df(df, column):
    """
    Remove outliers de uma coluna específica de um DataFrame usando o intervalo interquartil (IQR).

    Args:
    df (DataFrame): DataFrame do qual os outliers serão removidos.
    column (str): Nome da coluna na qual os outliers serão identificados e removidos.

    Returns:
    DataFrame: DataFrame sem os outliers identificados na coluna especificada.
    """
    Q1 = df[column].quantile(0.25)
    Q3 = df[column].quantile(0.75)
    IQR = Q3 - Q1
    upper_bound = Q3 + 1.5 * IQR
    return df[df[column] <= upper_bound]


# 3. FUNÇÕES DE INTERFACE E FILTROS


def add_custom_style():
    st.markdown(
        """
        <style>
        .filter-status {
            padding: 10px;
            border-radius: 5px;
            margin: 5px 0;
            background-color: #f0f2f6;
        }
        .filter-badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 10px;
            background-color: #e6e6e6;
            margin: 2px;
        }
        </style>
    """,
        unsafe_allow_html=True,
    )


@st.cache_data
def get_filter_options_dependent(
    df: pd.DataFrame,
    column_name: str,
    start_date: datetime.date,
    end_date: datetime.date,
    previous_filters: Optional[Dict[str, List[str]]] = None,
) -> List[str]:
    """
    Obtém opções únicas de uma coluna considerando período e filtros anteriores

    Args:
        df (pd.DataFrame): DataFrame fonte dos dados
        column_name (str): Nome da coluna para obter opções
        start_date (datetime.date): Data inicial do período
        end_date (datetime.date): Data final do período
        previous_filters (dict, optional): Dicionário com filtros anteriores

    Returns:
        list: Lista de opções únicas incluindo 'Todos'
    """
    # Cria uma cópia do DataFrame para não modificar o original
    df_filtered = df.copy()

    # Primeiro aplica o filtro de data
    start_datetime = pd.to_datetime(start_date)
    end_datetime = pd.to_datetime(end_date)

    # Se start_date e end_date são iguais, ajustar end_datetime para incluir todo o dia
    if start_date == end_date:
        end_datetime = end_datetime + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    df_filtered = df_filtered[
        (df_filtered[config["excel_columns"]["date"]] >= start_datetime)
        & (df_filtered[config["excel_columns"]["date"]] <= end_datetime)
    ]

    # Depois aplica os filtros anteriores se existirem
    if previous_filters:
        for col, values in previous_filters.items():
            if values and "Todos" not in values:
                df_filtered = df_filtered[df_filtered[col].isin(values)]

    # Obtém valores únicos e ordenados
    unique_values = sorted(df_filtered[column_name].unique().tolist())

    return ["Todos"] + unique_values


def create_dependent_multiselect(
    df, column_name, label, start_date, end_date, previous_filters=None, key=None
):
    """
    Cria um filtro multiselect que depende do período e das seleções anteriores

    Args:
        df (DataFrame): DataFrame fonte
        column_name (str): Nome da coluna para filtrar
        label (str): Texto do label do multiselect
        start_date (datetime.date): Data inicial do período
        end_date (datetime.date): Data final do período
        previous_filters (dict): Dicionário com filtros anteriores
        key (str): Chave única para o componente Streamlit

    Returns:
        list: Valores selecionados
    """
    options = get_filter_options_dependent(
        df, column_name, start_date, end_date, previous_filters
    )

    # Se não houver opções além de "Todos", desabilita o filtro
    if len(options) <= 1:
        st.warning(f"Nenhuma opção disponível para {label} no período selecionado")
        return ["Todos"]

    selected = st.multiselect(label, options=options, default=["Todos"], key=key)

    # Se nada selecionado, considera todos selecionados
    if not selected:
        return options[1:]  # Retorna todos exceto "Todos"

    # Se "Todos" está entre os selecionados, retorna todos
    if "Todos" in selected:
        return options[1:]

    return selected


def show_filter_status(operadores, alimentos, dietas, start_date, end_date):
    """
    Mostra o status atual dos filtros aplicados incluindo o período

    Args:
        operadores (list): Lista de operadores selecionados
        alimentos (list): Lista de alimentos selecionados
        dietas (list): Lista de dietas selecionadas
        start_date (datetime.date): Data inicial
        end_date (datetime.date): Data final
    """
    st.sidebar.markdown("### Filtros Ativos")

    # Mostrar período selecionado
    if start_date == end_date:
        st.sidebar.write(f"**Período:** {start_date.strftime('%d/%m/%Y')}")
    else:
        st.sidebar.write(
            f"**Período:** {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        )

    filtros = {"Operadores": operadores, "Alimentos": alimentos, "Dietas": dietas}

    for nome, valores in filtros.items():
        if valores and "Todos" not in valores:
            st.sidebar.write(f"**{nome}:** {', '.join(valores)}")
        else:
            st.sidebar.write(f"**{nome}:** Todos")


def add_reset_filters_button():
    """
    Adiciona um botão para resetar todos os filtros

    Returns:
        bool: True se o botão foi clicado
    """
    if st.sidebar.button("🔄 Limpar Filtros"):
        # Resetar todos os filtros para "Todos"
        for key in ["operators", "foods", "diets"]:
            if key in st.session_state:
                st.session_state[key] = ["Todos"]
        return True
    return False


def validate_selections(df_filtered, df_original, start_date, end_date):
    """
    Valida os resultados da filtragem e mostra feedback

    Args:
        df_filtered (DataFrame): DataFrame após aplicação dos filtros
        df_original (DataFrame): DataFrame original antes dos filtros
        start_date (datetime.date): Data inicial selecionada
        end_date (datetime.date): Data final selecionada

    Returns:
        bool: True se a validação passou, False caso contrário
    """
    # Primeiro filtra o df_original por data para ter uma base de comparação correta
    start_datetime = pd.to_datetime(start_date)
    end_datetime = pd.to_datetime(end_date)
    if start_date == end_date:
        end_datetime = end_datetime + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    df_period = df_original[
        (df_original[config["excel_columns"]["date"]] >= start_datetime)
        & (df_original[config["excel_columns"]["date"]] <= end_datetime)
    ]

    total_periodo = len(df_period)
    registros_filtrados = len(df_filtered)

    if registros_filtrados == 0:
        st.warning("⚠️ Os filtros selecionados não retornaram resultados para o período")
        return False

    if registros_filtrados < total_periodo:
        reducao = ((total_periodo - registros_filtrados) / total_periodo) * 100
        st.info(
            f"ℹ️ Filtros ativos: {registros_filtrados:,} de {total_periodo:,} registros do período ({reducao:.1f}% redução)"
        )

    return True


def flexible_date_selection(df, date_column):
    """
    Permite ao usuário selecionar uma única data ou um intervalo de datas.

    Args:
    df (DataFrame): DataFrame contendo os dados.
    date_column (str): Nome da coluna de data no DataFrame.

    Returns:
    tuple: (start_date, end_date)
    """
    min_date = df[date_column].min().date()
    max_date = df[date_column].max().date()

    selection_type = st.radio(
        "Selecione o tipo de filtro de data:", ["Data única", "Intervalo de datas"]
    )

    if selection_type == "Data única":
        selected_date = st.date_input(
            "Selecione a data:", min_value=min_date, max_value=max_date, value=min_date
        )
        start_date = end_date = selected_date
    else:
        start_date, end_date = st.date_input(
            "Selecione o intervalo de datas:",
            [min_date, max_date],
            min_value=min_date,
            max_value=max_date,
        )

        # Ajusta end_date para incluir todo o último dia
        end_date = end_date + timedelta(days=1) - timedelta(seconds=1)

    return start_date, end_date


def filter_data(df, operadores, alimentos, dietas, start_date, end_date):
    """
    Filtra os dados com base nos operadores, alimentos, dietas e intervalo de datas fornecidos.

    Args:
    df (DataFrame): DataFrame contendo os dados brutos.
    operadores (list): Lista de operadores selecionados para filtrar os dados.
    alimentos (list): Lista de alimentos selecionados para filtrar os dados.
    dietas (list): Lista de dietas selecionadas para filtrar os dados.
    start_date (datetime.date): Data de início do filtro de datas.
    end_date (datetime.date): Data de término do filtro de datas.

    Returns:
    DataFrame: DataFrame filtrado conforme os critérios especificados.
    """
    # Converter start_date e end_date para datetime
    start_datetime = pd.to_datetime(start_date)
    end_datetime = pd.to_datetime(end_date)

    # Se start_date e end_date são iguais, ajustar end_datetime para incluir todo o dia
    if start_date == end_date:
        end_datetime = end_datetime + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)

    # Filtrar por data
    df = df[
        (df[config["excel_columns"]["date"]] >= start_datetime)
        & (df[config["excel_columns"]["date"]] <= end_datetime)
    ]

    # Filtrar por operadores
    if "Todos" not in operadores:
        df = df[df[config["excel_columns"]["operator"]].isin(operadores)]

    # Filtrar por alimentos
    if "Todos" not in alimentos:
        df = df[df[config["excel_columns"]["alimento"]].isin(alimentos)]

    # Filtrar por dietas
    if "Todos" not in dietas:
        df = df[df[config["excel_columns"]["nome"]].isin(dietas)]

    return df


# 4. FUNÇÕES DE VISUALIZAÇÃO


def create_histogram(
    df, start_date, end_date, remove_outliers, pesos_relativos, config
):
    """
    Cria e exibe um histograma com base nas médias ponderadas das diferenças percentuais.

    Args:
    df (DataFrame): DataFrame contendo as médias ponderadas das diferenças percentuais.
    start_date (datetime): Data de início para o filtro de datas.
    end_date (datetime): Data de término para o filtro de datas.
    remove_outliers (bool): Se True, remove outliers dos dados antes de criar o histograma.
    pesos_relativos (dict): Dicionário contendo os pesos relativos de cada tipo de alimento.
    config (dict): Dicionário de configuração contendo as informações necessárias.

    Returns:
    matplotlib.figure.Figure: A figura contendo o histograma gerado.
    """
    figsize = tuple(config["visualization"]["histogram_figsize"])
    fig, ax = plt.subplots(figsize=figsize)

    if remove_outliers:
        df = remove_outliers_from_df(df, "MÉDIA PONDERADA (%)")

    data = df["MÉDIA PONDERADA (%)"]
    data = data[data >= 0]

    lower_bound, upper_bound, n_bins = calculate_histogram_bins(data)

    n, bins, patches = ax.hist(
        data, bins=n_bins, range=(lower_bound, upper_bound), edgecolor="black"
    )

    color_histogram_bars(patches, bins)

    ax.set_xlabel(config["visualization"]["x_label"])
    ax.set_ylabel(config["visualization"]["y_label"])
    ax.set_title(config["visualization"]["histogram_title"])

    tolerance = config["analysis"]["tolerance_threshold"]
    vertical_line_config = config["visualization"]["vertical_line"]
    ax.axvline(
        x=tolerance,
        color=vertical_line_config["color"],
        linestyle=vertical_line_config["linestyle"],
        linewidth=vertical_line_config["linewidth"],
        label=f"{vertical_line_config['label']} ({tolerance}%)",
    )

    legend_config = config["visualization"]["legend"]
    ax.legend(
        loc=legend_config["location"],
        fontsize=legend_config["fontsize"],
        frameon=legend_config["frameon"],
        facecolor=legend_config["facecolor"],
        edgecolor=legend_config["edgecolor"],
        fancybox=legend_config["fancybox"],
        framealpha=legend_config["framealpha"],
        bbox_to_anchor=legend_config["bbox_to_anchor"],
    )

    grid_config = config["visualization"]["grid_style"]
    ax.grid(
        axis=grid_config["axis"],
        linestyle=grid_config["linestyle"],
        linewidth=grid_config["linewidth"],
    )
    ax.set_axisbelow(True)

    ax.xaxis.set_major_locator(MaxNLocator(integer=True, nbins=20, prune="both"))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda x, p: f"{int(x)}"))

    brasilia_tz = pytz.timezone(config["timezone"])
    now_brasilia = datetime.datetime.now(brasilia_tz)

    footer_config = config["visualization"]["footer"]
    footer_texts = footer_config["texts"]
    footer_positions = footer_config["positions"]
    footer_alignments = footer_config["alignments"]

    plt.figtext(
        footer_positions["period_position"][0],
        footer_positions["period_position"][1],
        footer_texts["period_text"].format(
            start_date=start_date.strftime("%d/%m/%Y"),
            end_date=end_date.strftime("%d/%m/%Y"),
        ),
        ha=footer_alignments["period_alignment"],
        fontsize=footer_config["fontsize"],
    )
    plt.figtext(
        footer_positions["total_position"][0],
        footer_positions["total_position"][1],
        footer_texts["total_text"].format(total_batidas=len(df)),
        ha=footer_alignments["total_alignment"],
        fontsize=footer_config["fontsize"],
    )
    plt.figtext(
        footer_positions["generated_position"][0],
        footer_positions["generated_position"][1],
        footer_texts["generated_text"].format(
            generated_time=now_brasilia.strftime("%d/%m/%Y %H:%M")
        ),
        ha=footer_alignments["generated_alignment"],
        fontsize=footer_config["fontsize"],
    )

    weights_config = config["visualization"]["weights_table"]
    pesos_text = "Pesos relativos dos tipos de alimentos:\n"
    pesos_text += "\n".join(
        [f"{tipo:>20}: {peso:>4.1f}" for tipo, peso in pesos_relativos.items()]
    )
    fig.text(
        weights_config["position"][0],
        weights_config["position"][1],
        pesos_text,
        ha="right",
        fontsize=weights_config["fontsize"],
        va="top",
        linespacing=1.5,
        bbox=dict(
            facecolor=weights_config["facecolor"],
            alpha=weights_config["alpha"],
            boxstyle=weights_config["boxstyle"],
        ),
    )

    subplot_config = config["visualization"]["subplot_adjust"]
    plt.subplots_adjust(
        left=subplot_config["left"],
        right=subplot_config["right"],
        bottom=subplot_config["bottom"],
        top=subplot_config["top"],
    )

    return fig


def calculate_histogram_bins(data):
    """
    Calcula os limites e o número de bins para o histograma usando a regra de Freedman-Diaconis.

    Args:
    data (Series): Série ou lista de dados numéricos para os quais o histograma será calculado.

    Returns:
    tuple: (limite inferior, limite superior, número de bins) para o histograma.
    """
    q1, q3 = np.percentile(data, [25, 75])
    iqr = q3 - q1

    lower_bound = data[data >= 0].min()
    upper_bound = q3 + 1.5 * iqr

    upper_bound = np.ceil(upper_bound)

    bin_width = 2 * iqr * (len(data) ** (-1 / 3))
    n_bins = int((upper_bound - lower_bound) / bin_width)
    n_bins = min(n_bins, 100)

    return lower_bound, upper_bound, n_bins


def color_histogram_bars(patches, bins):
    """
    Aplica cores às barras do histograma com base nos valores dos bins.

    Args:
    patches (list): Lista de patches (barras) do histograma.
    bins (list): Lista de valores dos bins.
    """
    for patch, bin_value in zip(patches, bins[:-1]):
        if bin_value >= 3:
            color_intensity = min((bin_value - 3) / (bins[-1] - 3), 1)
            patch.set_facecolor((1, 0, 0, color_intensity))
        else:
            color_intensity = min((3 - bin_value) / 3, 1)
            patch.set_facecolor((0, 1, 0, color_intensity))


def create_statistics_dataframe(
    weighted_average_df, remove_outliers=False, config=None
):
    """
    Cria um DataFrame com estatísticas das diferenças percentuais.

    Argumentos:
    weighted_average_df: DataFrame com médias ponderadas das diferenças percentuais.
    remove_outliers: Se True, remove outliers antes dos cálculos.
    config: Dicionário com parâmetros de configuração.

    Retorna:
    DataFrame com estatísticas.
    """
    if config is None:
        raise ValueError(
            "O parâmetro `config` não foi fornecido para `create_statistics_dataframe`."
        )

    if "statistics" not in config or "interval_limits" not in config["statistics"]:
        raise KeyError(
            "A chave 'statistics' ou 'interval_limits' não está presente no `config`. "
            "Verifique se o arquivo de configuração está correto e atualizado."
        )

    df = weighted_average_df.copy()

    if remove_outliers:
        df = remove_outliers_from_df(df, "MÉDIA PONDERADA (%)")

    interval_limits = config["statistics"]["interval_limits"]

    stats_data = {
        "Estatística": [
            "Número de Batidas",
            "Média Ponderada (%)",
            "Mediana Ponderada (%)",
            f"Diferença entre {interval_limits['low_1']}% e {interval_limits['high_1']}%",
            f"Diferença entre {interval_limits['low_2']}% e {interval_limits['high_2']}%",
            f"Diferença acima de {interval_limits['high_2']}%",
        ],
        "Valor": [
            len(df),
            round(df["MÉDIA PONDERADA (%)"].mean(), 1),
            round(df["MÉDIA PONDERADA (%)"].median(), 1),
            (
                (df["MÉDIA PONDERADA (%)"] >= interval_limits["low_1"])
                & (df["MÉDIA PONDERADA (%)"] < interval_limits["high_1"])
            ).sum(),
            (
                (df["MÉDIA PONDERADA (%)"] >= interval_limits["low_2"])
                & (df["MÉDIA PONDERADA (%)"] < interval_limits["high_2"])
            ).sum(),
            (df["MÉDIA PONDERADA (%)"] >= interval_limits["high_2"]).sum(),
        ],
    }

    stats_df = pd.DataFrame(stats_data)

    return stats_df


# 5. FUNÇÕES DE EXPORTAÇÃO


def save_histogram_as_image(fig):
    """
    Salva o histograma gerado como uma imagem PNG e retorna um link para download.

    Args:
    fig (matplotlib.figure.Figure): A figura do histograma a ser salva.

    Returns:
    str: Link para download da imagem gerada.
    """
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=300, bbox_inches="tight")
    buf.seek(0)
    b64 = base64.b64encode(buf.getvalue()).decode()
    href = (
        f'<a href="data:image/png;base64,{b64}" '
        f'download="histograma.png">- histograma(png)</a>'
    )
    return href


def save_statistics_as_csv(stats_df):
    """
    Salva as estatísticas geradas como um arquivo CSV e retorna um link para download.

    Args:
    stats_df (DataFrame): DataFrame contendo as estatísticas calculadas.

    Returns:
    str: Link para download do arquivo CSV gerado.
    """
    csv = stats_df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = (
        f'<a href="data:file/csv;base64,{b64}" '
        f'download="estatisticas.csv">- estatísticas (csv)</a>'
    )
    return href


def save_dataframe_as_excel(df):
    """
    Salva o DataFrame processado final como um arquivo Excel e retorna um buffer para download.

    Args:
    df (DataFrame): DataFrame contendo os dados processados e organizados por BATIDAS e suas médias de diferenças.

    Returns:
    BytesIO: Um buffer contendo o arquivo Excel gerado.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados Processados")
        workbook = writer.book
        worksheet = writer.sheets["Dados Processados"]

        color_config = config["ui"]["conditional_formatting"]["colors"]

        green_fill = PatternFill(
            start_color=color_config["green"],
            end_color=color_config["green"],
            fill_type="solid",
        )
        red_fill_light = PatternFill(
            start_color=color_config["red_light"],
            end_color=color_config["red_light"],
            fill_type="solid",
        )
        red_fill_intense = PatternFill(
            start_color=color_config["red_intense"],
            end_color=color_config["red_intense"],
            fill_type="solid",
        )
        black_fill = PatternFill(
            start_color=color_config["black"],
            end_color=color_config["black"],
            fill_type="solid",
        )
        white_font = openpyxl.styles.Font(color=color_config["white_font"])

        for row in range(2, len(df) + 2):
            cell = worksheet[f"B{row}"]
            cell.number_format = "0.00"
            if cell.value <= config["statistics"]["interval_limits"]["low_1"]:
                cell.fill = green_fill
            elif (
                config["statistics"]["interval_limits"]["low_1"]
                < cell.value
                <= config["statistics"]["interval_limits"]["high_1"]
            ):
                cell.fill = red_fill_light
            elif (
                config["statistics"]["interval_limits"]["low_2"]
                < cell.value
                <= config["statistics"]["interval_limits"]["high_2"]
            ):
                cell.fill = red_fill_intense
            elif cell.value > config["statistics"]["interval_limits"]["high_2"]:
                cell.fill = black_fill
                cell.font = white_font

    buffer.seek(0)
    return buffer


# 6. FUNÇÃO PRINCIPAL


def main():
    """
    Função principal que controla a execução do programa e a interação com o usuário via Streamlit.
    """

    add_custom_style()

    st.title(config["ui"]["page_title"])

    col1, col2 = st.columns([1, 3])

    with col1:
        st.header(config["ui"]["analysis_config_header"])

        uploaded_file = st.file_uploader(
            config["ui"]["file_uploader"]["label"],
            type=config["ui"]["file_uploader"]["allowed_types"],
        )

        if uploaded_file is None:
            st.warning("selecione o arquivo .XLSX para análise.")
            return

        df = load_and_process_data(uploaded_file)

        if df is not None:
            st.success("Arquivo carregado com sucesso!")

            # Adicionar botão de reset na sidebar
            if add_reset_filters_button():
                st.rerun()

            # Seleção de datas (agora é o primeiro filtro)
            min_date = df[config["excel_columns"]["date"]].min().date()
            max_date = df[config["excel_columns"]["date"]].max().date()
            start_date, end_date = flexible_date_selection(
                df, config["excel_columns"]["date"]
            )

            # Inicializa dicionário de filtros
            filters_dict = {}

            # 1. Filtro de Operadores (agora considera o período)
            operadores_selecionados = create_dependent_multiselect(
                df,
                config["excel_columns"]["operator"],
                config["ui"]["multiselect"]["operator_label"],
                start_date,
                end_date,
                key="operators",
            )
            filters_dict[config["excel_columns"]["operator"]] = operadores_selecionados

            # 2. Filtro de Alimentos - dependente do período e operadores
            alimentos_selecionados = create_dependent_multiselect(
                df,
                config["excel_columns"]["alimento"],
                config["ui"]["multiselect"]["food_label"],
                start_date,
                end_date,
                previous_filters=filters_dict,
                key="foods",
            )
            filters_dict[config["excel_columns"]["alimento"]] = alimentos_selecionados

            # 3. Filtro de Dietas - dependente do período, operadores e alimentos
            dietas_selecionadas = create_dependent_multiselect(
                df,
                config["excel_columns"]["nome"],
                config["ui"]["multiselect"]["diet_label"],
                start_date,
                end_date,
                previous_filters=filters_dict,
                key="diets",
            )
            filters_dict[config["excel_columns"]["nome"]] = dietas_selecionadas

            # Mostrar status dos filtros na sidebar
            show_filter_status(
                operadores_selecionados,
                alimentos_selecionados,
                dietas_selecionadas,
                start_date,
                end_date,
            )

            # Aplicar filtros
            df_filtered = filter_data(
                df,
                operadores_selecionados,
                alimentos_selecionados,
                dietas_selecionadas,
                start_date,
                end_date,
            )

            # Validar resultado da filtragem
            if not validate_selections(df_filtered, df, start_date, end_date):
                return

            st.success("Filtro aplicado com sucesso!")

            st.subheader(config["ui"]["food_weights_subheader"])
            tipos_alimentos = df[config["excel_columns"]["tipo"]].unique().tolist()
            pesos_relativos = {
                tipo: st.slider(
                    f"Peso para tipo de alimento '{tipo}':",
                    min_value=config["slider"]["min_value"],
                    max_value=config["slider"]["max_value"],
                    value=config["slider"]["default_value"],
                    step=config["slider"]["step"],
                )
                for tipo in tipos_alimentos
            }

            remover_outliers = st.checkbox(
                config["ui"]["remove_outliers"]["label"],
                help=config["ui"]["remove_outliers"]["help"],
            )

            iniciar_analise = st.button(config["ui"]["generate_button"])

    with col2:
        if df is not None and iniciar_analise:
            st.header(config["ui"]["results_header"])

            weighted_average_df = calculate_weighted_average_with_weights(
                df_filtered, pesos_relativos, config
            )

            if weighted_average_df is not None:
                fig = create_histogram(
                    weighted_average_df,
                    start_date,
                    end_date,
                    remover_outliers,
                    pesos_relativos,
                    config=config,
                )
                st.pyplot(fig)

                col1, col2, col3 = st.columns(3)

                with col1:
                    st.subheader(config["ui"]["statistics_title"])
                    stats_df = create_statistics_dataframe(
                        weighted_average_df, remover_outliers, config=config
                    )
                    st.write(stats_df)

                with col2:
                    st.subheader(config["ui"]["food_weights_subheader"])
                    pesos_df = pd.DataFrame(
                        list(pesos_relativos.items()),
                        columns=["Tipo de Alimento", "Peso Relativo"],
                    )
                    st.write(pesos_df)

                with col3:
                    st.subheader("downloads")

                    buf = io.BytesIO()
                    fig.savefig(buf, format="png", dpi=300, bbox_inches="tight")
                    buf.seek(0)
                    st.download_button(
                        label="histograma (PNG)",
                        data=buf,
                        file_name="histograma.png",
                        mime="image/png",
                    )

                    csv_data = stats_df.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        label="estatísticas (CSV)",
                        data=csv_data,
                        file_name="estatisticas.csv",
                        mime="text/csv",
                    )

                    excel_buffer = save_dataframe_as_excel(weighted_average_df)
                    st.download_button(
                        label="dados processados (Excel)",
                        data=excel_buffer,
                        file_name="dados_processados.xlsx",
                        mime=(
                            "application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet"
                        ),
                    )

                    if remover_outliers:
                        st.info(config["ui"].get("outliers_removed_message"))

            else:
                st.error(
                    "Não foi possível calcular as médias ponderadas. "
                    "Verifique os dados e tente novamente."
                )


# 7. ENTRY POINT
if __name__ == "__main__":
    main()
